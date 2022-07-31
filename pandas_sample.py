import openpyxl
import networkx as nx
import pandas as pd
import matplotlib.pyplot as plt


def read_from_file():
    print('started')
    study_name_list = list()
    treatment_name_list = list()
    work_book = openpyxl.load_workbook('another_data_250.xlsx')
    worksheet = work_book.active
    total_row = worksheet.max_row
    for i in range(2, total_row+1):
        study_name = worksheet.cell(row=i, column=1).value
        treatment_name = worksheet.cell(row=i, column=2).value
        study_name_list.append(study_name)
        treatment_name_list.append(treatment_name)

    study_treatment_dict = dict()
    data_len = len(study_name_list)
    for i in range(data_len):
        existing_treatments = study_treatment_dict.get(study_name_list[i], list())
        existing_treatments.append(treatment_name_list[i].strip())
        study_treatment_dict[study_name_list[i].strip()] = existing_treatments
    pahela_list = list()
    doosra_list = list()
    for key, value in study_treatment_dict.items():
        treatment_list_size = len(value)
        if treatment_list_size > 1:
            for i in range(treatment_list_size):
                for j in range(i + 1, treatment_list_size):
                    pahela_list.append(value[i])
                    doosra_list.append(value[j])
        elif treatment_list_size == 1:
            pahela_list.append(value[0])
            doosra_list.append('')

    my_data_frame = {
        'from': pahela_list,
        'to': doosra_list
    }
    my_data_frame = pd.DataFrame(my_data_frame)
    my_data_frame['weight'] = (my_data_frame.groupby(['from', 'to'])['from'].transform('size'))/10
    graph_obj = nx.from_pandas_edgelist(my_data_frame, source="from", target="to", create_using=nx.Graph(), edge_attr='weight')
    graph_obj.remove_node('')

    weights = [graph_obj[u][v]['weight'] for u, v in graph_obj.edges()]
    edge_weights = nx.get_edge_attributes(graph_obj, "weight")
    for key in edge_weights:
        edge_weights[key] = int(edge_weights[key] * 10)
    layout = nx.shell_layout(graph_obj, scale=2)
    node_degree = dict(nx.degree(graph_obj))
    for key, item in node_degree.items():
        if item == 0:
            node_degree[key] = 1
    nx.draw(graph_obj, with_labels=True, pos=layout, font_size=8, width=weights, node_size=[v * 100 for v in node_degree.values()], node_color='red')

    nx.draw_networkx_edge_labels(graph_obj, pos=layout, edge_labels=edge_weights, font_size=6)
    plt.show()


if __name__ == '__main__':
    read_from_file()
