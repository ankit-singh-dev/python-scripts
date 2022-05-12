import os.path
import time
import requests
from pathlib import Path
import openpyxl
from bs4 import BeautifulSoup


def extract_pdf(data_from_xlsx_file):
    # time_pass_url = "https://vk.com/js/api/openapi.js?49"
    # requests.get(time_pass_url)
    base_url = "https://sci-hub.st/"
    # requests.get(base_url)

    success_failure_dict = dict()
    for key, value in data_from_xlsx_file.items():
        hit_url = base_url + value[0]
        print(hit_url)
        try:
            hit_response = requests.get(hit_url)
            time.sleep(3)
            soup = BeautifulSoup(hit_response.text, 'html.parser')
            file_url_res = soup.find('embed').attrs["src"]
            if 'https' not in file_url_res:
                file_url_res = 'https:' + soup.find('embed').attrs["src"]

            print(file_url_res)
            file_res = requests.get(file_url_res)
            filename = Path(key + '.pdf')
            filename.write_bytes(file_res.content)
            print('file downloaded successfully')
            success_failure_dict[key] = "Yes"
        except:
            success_failure_dict[key] = "No"
            print('download failed', key)

    return success_failure_dict


def read_from_xlsx_file(xlsx_file_path):
    print("Reading from xlsx file")
    # xlsx_file_path = "C:\\Users\\Ankit_Singh1\\Downloads\\Excel file.xlsx"
    wb_obj = openpyxl.load_workbook(xlsx_file_path)
    sheet_obj = wb_obj.active
    file_name_pm_ids_dict = dict()
    for i in range(2, sheet_obj.max_row+1):
        file_name = sheet_obj.cell(row=i, column=1).value
        doi_value = sheet_obj.cell(row=i, column=2).value
        pm_id = sheet_obj.cell(row=i, column=3).value
        file_name_pm_ids_dict[file_name] = doi_value,pm_id

    print("Reading from xlsx file completed")
    return file_name_pm_ids_dict


def populate_success_failure(success_failure_status_dict, xlsx_file_path):
    print("Writing to existing xlsx file")
    # xlsx_file_path = "C:\\Users\\Ankit_Singh1\\Downloads\\Excel file.xlsx"
    str_file_path = str(xlsx_file_path)
    file_components = str_file_path.split(".")
    new_file_name = file_components[0] + "_Result." + file_components[1]
    print(new_file_name)
    wb_obj = openpyxl.load_workbook(xlsx_file_path)
    sheet_obj = wb_obj.active
    for i in range(2, sheet_obj.max_row+1):
        sheet_obj.cell(row=i, column=4).value = success_failure_status_dict.get(sheet_obj.cell(row=i, column=1).value)
    # wb_obj.save("C:\\Users\\Ankit_Singh1\\Downloads\\ResultFile.xlsx")
    wb_obj.save(new_file_name)


if __name__ == '__main__':
    input_file_path = input("Enter xlsx file path\n")
    if os.path.isfile(input_file_path):
        right_path = Path(input_file_path)
        get_data_from_xlsx_file_dict = read_from_xlsx_file(right_path)
        success_failure_status = extract_pdf(get_data_from_xlsx_file_dict)
        populate_success_failure(success_failure_status, right_path)
    else:
        print("File doesn't exist")